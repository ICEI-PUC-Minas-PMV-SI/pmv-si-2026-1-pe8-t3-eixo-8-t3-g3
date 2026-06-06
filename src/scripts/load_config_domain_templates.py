#!/usr/bin/env python3
"""
Load governed config/domain Excel templates into PostgreSQL.

Default behavior is safe: validate templates and write a generated SQL file.
Use --execute to run the generated SQL with psql.
"""

from __future__ import annotations

import argparse
import getpass
import subprocess
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_DIR = REPO_ROOT / "artefatos" / "config_templates"
DEFAULT_DDL = REPO_ROOT / "bd" / "config" / "config_domain_template_ddl.sql"
DEFAULT_SQL_OUT = REPO_ROOT / "artefatos" / "generated" / "config_domain_template_load.sql"


TEMPLATE_FILES = {
    "agents": "brokerlab_config_agents.xlsx",
    "targets": "brokerlab_config_targets.xlsx",
    "finance": "brokerlab_domain_finance.xlsx",
    "crm": "brokerlab_domain_crm.xlsx",
    "assets": "brokerlab_config_assets.xlsx",
    "calendar": "brokerlab_config_calendar.xlsx",
}


@dataclass
class ValidationIssue:
    severity: str
    source_file: str
    sheet_name: str
    row_number: int | None
    field_name: str | None
    message: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )
    return " ".join(text.split())


def sql_string(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return "'" + value.isoformat().replace("'", "''") + "'"
    text = str(value)
    if text == "":
        return "NULL"
    return "'" + text.replace("'", "''") + "'"


def sql_bool(value: Any, default: bool | None = None) -> str:
    parsed = parse_bool(value, default=default)
    if parsed is None:
        return "NULL"
    return "TRUE" if parsed else "FALSE"


def parse_bool(value: Any, default: bool | None = None) -> bool | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "t", "yes", "y", "1", "sim"}:
        return True
    if text in {"false", "f", "no", "n", "0", "nao", "não"}:
        return False
    return default


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def parse_int(value: Any) -> int | None:
    dec = parse_decimal(value)
    if dec is None:
        return None
    return int(dec)


def sql_number(value: Any, default: Any = None) -> str:
    if value is None or value == "":
        value = default
    if value is None or value == "":
        return "NULL"
    dec = parse_decimal(value)
    if dec is None:
        return "NULL"
    return format(dec, "f")


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and value != "" for value in values):
            continue
        item = dict(zip(headers, values))
        item["_row_number"] = row_number
        rows.append(item)
    return rows


def require_files(template_dir: Path) -> dict[str, Path]:
    paths = {key: template_dir / filename for key, filename in TEMPLATE_FILES.items()}
    missing = [str(path) for key, path in paths.items() if key != "calendar" and not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required template files: " + ", ".join(missing))
    return paths


def load_templates(template_dir: Path) -> dict[str, list[dict[str, Any]]]:
    paths = require_files(template_dir)
    return {
        "agent_profile": read_sheet(paths["agents"], "agent_profile"),
        "agent_alias": read_sheet(paths["agents"], "agent_alias"),
        "agent_target_month": read_sheet(paths["targets"], "agent_target_month"),
        "payment_method": read_sheet(paths["finance"], "payment_method"),
        "currency_mapping": read_sheet(paths["finance"], "currency_mapping"),
        "lead_status_mapping": read_sheet(paths["crm"], "lead_status_mapping"),
        "asset_catalog_override": read_sheet(paths["assets"], "asset_catalog_override"),
        "holiday_exception": read_sheet(paths["calendar"], "holiday_exception") if paths["calendar"].exists() else [],
    }


def issue(
    issues: list[ValidationIssue],
    severity: str,
    source_file: str,
    sheet_name: str,
    row: dict[str, Any] | None,
    field_name: str | None,
    message: str,
) -> None:
    issues.append(
        ValidationIssue(
            severity=severity,
            source_file=source_file,
            sheet_name=sheet_name,
            row_number=row.get("_row_number") if row else None,
            field_name=field_name,
            message=message,
        )
    )


def validate(data: dict[str, list[dict[str, Any]]]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []

    agent_keys: set[str] = set()
    active_agent_names: set[str] = set()
    alias_pairs: set[tuple[str, str]] = set()

    for row in data["agent_profile"]:
        agent_key = row.get("agent_key")
        agent_name = row.get("agent_name")
        team_name = row.get("team_name")
        agent_type = row.get("agent_type")
        agent_level = row.get("agent_level")
        is_active = parse_bool(row.get("is_active"))

        if not agent_key:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_key", "agent_key is required")
            continue
        if agent_key in agent_keys:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_key", f"duplicate agent_key: {agent_key}")
        agent_keys.add(str(agent_key))

        if not agent_name:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_name", "agent_name is required")
        if not team_name:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "team_name", "team_name is required by config.agent_profile")
        if agent_type not in {"individual", "pool", "system"}:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_type", f"invalid agent_type: {agent_type}")
        if agent_level not in {None, "", "trainee", "inter", "pro"}:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_level", f"invalid agent_level: {agent_level}")
        if is_active is None:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "is_active", f"invalid is_active: {row.get('is_active')}")

        active_name = normalize_text(agent_name)
        if is_active and agent_type == "individual" and active_name:
            if active_name in active_agent_names:
                issue(issues, "error", TEMPLATE_FILES["agents"], "agent_profile", row, "agent_name", f"duplicate active individual agent_name: {agent_name}")
            active_agent_names.add(active_name)

    for row in data["agent_alias"]:
        agent_key = row.get("agent_key")
        alias_name = row.get("alias_name")
        source_system = row.get("source_system")
        if agent_key not in agent_keys:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_alias", row, "agent_key", f"agent_key not found in agent_profile: {agent_key}")
        if not alias_name:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_alias", row, "alias_name", "alias_name is required")
        if not source_system:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_alias", row, "source_system", "source_system is required")
        pair = (normalize_text(alias_name), str(source_system).strip().upper())
        if pair in alias_pairs:
            issue(issues, "error", TEMPLATE_FILES["agents"], "agent_alias", row, "alias_name", f"duplicate alias/source: {alias_name}/{source_system}")
        alias_pairs.add(pair)

    for row in data["agent_target_month"]:
        if not target_has_values(row):
            continue
        agent_key = row.get("agent_key")
        competence_month = parse_date(row.get("competence_month"))
        if agent_key not in agent_keys:
            issue(issues, "error", TEMPLATE_FILES["targets"], "agent_target_month", row, "agent_key", f"agent_key not found in agent_profile: {agent_key}")
        if not competence_month:
            issue(issues, "error", TEMPLATE_FILES["targets"], "agent_target_month", row, "competence_month", f"invalid competence_month: {row.get('competence_month')}")
        elif competence_month.day != 1:
            issue(issues, "error", TEMPLATE_FILES["targets"], "agent_target_month", row, "competence_month", "competence_month must be first day of month")
        for field in ("target_deposit_month_usd", "target_trade_day", "target_trade_month", "target_unique_month"):
            value = parse_decimal(row.get(field))
            if value is not None and value < 0:
                issue(issues, "error", TEMPLATE_FILES["targets"], "agent_target_month", row, field, "target cannot be negative")

    payment_codes: set[int] = set()
    for row in data["payment_method"]:
        code = parse_int(row.get("codigo"))
        if code is None:
            issue(issues, "error", TEMPLATE_FILES["finance"], "payment_method", row, "codigo", f"invalid codigo: {row.get('codigo')}")
            continue
        if code in payment_codes:
            issue(issues, "error", TEMPLATE_FILES["finance"], "payment_method", row, "codigo", f"duplicate codigo: {code}")
        payment_codes.add(code)
        if parse_bool(row.get("is_official")) is None:
            issue(issues, "error", TEMPLATE_FILES["finance"], "payment_method", row, "is_official", f"invalid is_official: {row.get('is_official')}")
        if parse_bool(row.get("is_active")) is None:
            issue(issues, "error", TEMPLATE_FILES["finance"], "payment_method", row, "is_active", f"invalid is_active: {row.get('is_active')}")

    lead_categories = {"Novo", "Em prospecção", "Perdido", "Inválido", "Operacional"}
    lead_codes: set[int] = set()
    for row in data["lead_status_mapping"]:
        code = parse_int(row.get("codigo"))
        if code is None:
            issue(issues, "error", TEMPLATE_FILES["crm"], "lead_status_mapping", row, "codigo", f"invalid codigo: {row.get('codigo')}")
            continue
        if code in lead_codes:
            issue(issues, "error", TEMPLATE_FILES["crm"], "lead_status_mapping", row, "codigo", f"duplicate codigo: {code}")
        lead_codes.add(code)
        if not row.get("descricao"):
            issue(issues, "error", TEMPLATE_FILES["crm"], "lead_status_mapping", row, "descricao", "descricao is required")
        if row.get("categoria") not in lead_categories:
            issue(issues, "error", TEMPLATE_FILES["crm"], "lead_status_mapping", row, "categoria", f"invalid categoria: {row.get('categoria')}")
        if parse_bool(row.get("eh_terminal")) is None:
            issue(issues, "error", TEMPLATE_FILES["crm"], "lead_status_mapping", row, "eh_terminal", f"invalid eh_terminal: {row.get('eh_terminal')}")

    asset_classes = {"forex", "commodity", "energy", "index", "stock", "crypto", "unknown"}
    asset_symbols: set[str] = set()
    for row in data["asset_catalog_override"]:
        symbol = row.get("sirix_symbol")
        if not symbol:
            issue(issues, "error", TEMPLATE_FILES["assets"], "asset_catalog_override", row, "sirix_symbol", "sirix_symbol is required")
            continue
        if symbol in asset_symbols:
            issue(issues, "error", TEMPLATE_FILES["assets"], "asset_catalog_override", row, "sirix_symbol", f"duplicate sirix_symbol: {symbol}")
        asset_symbols.add(str(symbol))
        if row.get("asset_class") not in asset_classes:
            issue(issues, "error", TEMPLATE_FILES["assets"], "asset_catalog_override", row, "asset_class", f"invalid asset_class: {row.get('asset_class')}")
        if parse_bool(row.get("is_major_asset")) is None:
            issue(issues, "error", TEMPLATE_FILES["assets"], "asset_catalog_override", row, "is_major_asset", f"invalid is_major_asset: {row.get('is_major_asset')}")
        if parse_bool(row.get("is_active")) is None:
            issue(issues, "error", TEMPLATE_FILES["assets"], "asset_catalog_override", row, "is_active", f"invalid is_active: {row.get('is_active')}")

    return issues


def target_has_values(row: dict[str, Any]) -> bool:
    fields = (
        "target_deposit_month_usd",
        "target_trade_day",
        "target_trade_month",
        "target_unique_month",
        "target_volume_month",
    )
    return any(row.get(field) not in (None, "") for field in fields)


def values_sql(rows: list[tuple[Any, ...]]) -> str:
    return ",\n".join("(" + ", ".join(str(value) for value in row) + ")" for row in rows)


def source_file(name: str) -> str:
    return TEMPLATE_FILES[name]


def generate_sql(data: dict[str, list[dict[str, Any]]], template_dir: Path, ddl_path: Path) -> str:
    ddl = ddl_path.read_text(encoding="utf-8")
    parts: list[str] = [
        ddl.rstrip(),
        "",
        "BEGIN;",
        "INSERT INTO config_import.load_batch (source_kind, template_dir, executed_by, completed_at, status, notes)",
        f"VALUES ('local_excel', {sql_string(str(template_dir))}, {sql_string(getpass.getuser())}, NOW(), 'completed', 'Loaded by scripts/load_config_domain_templates.py');",
        "",
    ]

    add_domain_sql(parts, data)
    add_config_sql(parts, data)

    parts.append("COMMIT;")
    parts.append("")
    return "\n".join(parts)


def add_domain_sql(parts: list[str], data: dict[str, list[dict[str, Any]]]) -> None:
    lead_rows = []
    for row in data["lead_status_mapping"]:
        lead_rows.append((
            sql_number(row.get("codigo")),
            sql_string(row.get("descricao")),
            sql_string(row.get("categoria")),
            sql_bool(row.get("eh_terminal")),
            sql_bool(row.get("is_active"), default=True),
            sql_string(source_file("crm")),
            sql_string(row.get("reviewed_by")),
            sql_string(row.get("notes")),
        ))
    if lead_rows:
        parts.append(
            "INSERT INTO domain.dom_lead_status "
            "(codigo, descricao, categoria, eh_terminal, is_active, source_file, reviewed_by, notes) VALUES\n"
            + values_sql(lead_rows)
            + ";"
        )

    payment_rows = []
    for row in data["payment_method"]:
        payment_rows.append((
            sql_number(row.get("codigo")),
            sql_string(row.get("official_name")),
            sql_string(row.get("payment_channel")),
            sql_bool(row.get("is_official"), default=False),
            sql_bool(row.get("is_active"), default=True),
            sql_string(source_file("finance")),
            sql_string(row.get("reviewed_by")),
            sql_string(row.get("notes")),
        ))
    if payment_rows:
        parts.append(
            "INSERT INTO domain.dom_payment_method "
            "(codigo, nome_inferido, payment_channel, eh_oficial, is_active, source_file, reviewed_by, notes) VALUES\n"
            + values_sql(payment_rows)
            + ";"
        )

    currency_rows = []
    for row in data["currency_mapping"]:
        currency_rows.append((
            sql_string(row.get("currency_guid")),
            sql_string(row.get("iso_code")),
            sql_string(row.get("currency_name")),
            sql_bool(row.get("is_active"), default=True),
            sql_string(source_file("finance")),
            sql_string(row.get("notes")),
        ))
    if currency_rows:
        parts.append(
            "INSERT INTO domain.dom_currency "
            "(currency_guid, iso_code, nome, is_active, source_file, notes) VALUES\n"
            + values_sql(currency_rows)
            + ";"
        )


def add_config_sql(parts: list[str], data: dict[str, list[dict[str, Any]]]) -> None:
    profile_rows = []
    for row in data["agent_profile"]:
        profile_rows.append((
            sql_string(row.get("agent_key")),
            sql_string(row.get("agent_name")),
            sql_string(row.get("agent_email")),
            sql_string(row.get("team_name")),
            sql_string(row.get("agent_level")),
            sql_string(row.get("seniority")),
            sql_string(row.get("agent_type")),
            sql_bool(row.get("is_active"), default=True),
            sql_string(parse_date(row.get("started_on"))),
            sql_string(parse_date(row.get("ended_on"))),
            sql_string(source_file("agents")),
            sql_string(row.get("notes")),
        ))
    if profile_rows:
        parts.append(
            "INSERT INTO config.agent_profile "
            "(agent_key, agent_name, agent_email, team_name, agent_level, seniority, agent_type, is_active, started_on, ended_on, source_file, notes) VALUES\n"
            + values_sql(profile_rows)
            + ";"
        )

    alias_selects = []
    for row in data["agent_alias"]:
        alias_selects.append(
            "SELECT ap.agent_id, {source_system}, {alias_name}, LOWER(TRIM(unaccent({alias_name}))), "
            "{is_primary}, {is_active}, {source_file}, {notes} "
            "FROM config.agent_profile ap WHERE ap.agent_key = {agent_key}".format(
                source_system=sql_string(row.get("source_system")),
                alias_name=sql_string(row.get("alias_name")),
                is_primary=sql_bool(row.get("is_primary"), default=False),
                is_active=sql_bool(row.get("is_active"), default=True),
                source_file=sql_string(source_file("agents")),
                notes=sql_string(row.get("notes")),
                agent_key=sql_string(row.get("agent_key")),
            )
        )
    if alias_selects:
        parts.append(
            "INSERT INTO config.agent_alias "
            "(agent_id, source_system, alias_name, normalized_alias, is_primary, is_active, source_file, notes)\n"
            + "\nUNION ALL\n".join(alias_selects)
            + ";"
        )

    parts.append(calendar_sql(data["holiday_exception"]))

    asset_rows = []
    for row in data["asset_catalog_override"]:
        asset_rows.append((
            sql_string(row.get("sirix_symbol")),
            sql_string(normalize_text(row.get("sirix_symbol"))),
            sql_string(row.get("display_name")),
            sql_string(row.get("asset_class") or "unknown"),
            sql_string(row.get("base_currency")),
            sql_string(row.get("quote_currency")),
            sql_number(row.get("contract_size")),
            sql_number(row.get("tick_size")),
            sql_number(row.get("tick_value")),
            sql_number(row.get("volume_multiplier")),
            sql_string(row.get("provider")),
            sql_string(row.get("provider_symbol")),
            sql_bool(row.get("is_major_asset"), default=False),
            sql_bool(row.get("is_active"), default=True),
            sql_string(row.get("classification_status") or "inferred"),
            sql_string(source_file("assets")),
            sql_string(row.get("notes")),
        ))
    if asset_rows:
        parts.append(
            "INSERT INTO config.asset_catalog "
            "(sirix_symbol, normalized_symbol, display_name, asset_class, base_currency, quote_currency, "
            "contract_size, tick_size, tick_value, volume_multiplier, provider, provider_symbol, "
            "is_major_asset, is_active, classification_status, source_file, notes) VALUES\n"
            + values_sql(asset_rows)
            + ";"
        )

    target_selects = []
    for row in data["agent_target_month"]:
        if not target_has_values(row):
            continue
        target_selects.append(
            "SELECT {month}::date, ap.agent_id, {deposit}, {trade_day}, {trade_month}::integer, {unique_month}, "
            "{volume_month}::numeric, {volume_unit}, 'template', {source_file}, {approved_by}, {is_active}, "
            "{created_by}, {notes} FROM config.agent_profile ap WHERE ap.agent_key = {agent_key}".format(
                month=sql_string(parse_date(row.get("competence_month"))),
                deposit=sql_number(row.get("target_deposit_month_usd"), default=0),
                trade_day=sql_number(row.get("target_trade_day"), default=0),
                trade_month=sql_number(row.get("target_trade_month")),
                unique_month=sql_number(row.get("target_unique_month"), default=0),
                volume_month=sql_number(row.get("target_volume_month")),
                volume_unit=sql_string(row.get("target_volume_unit")),
                source_file=sql_string(source_file("targets")),
                approved_by=sql_string(row.get("approved_by")),
                is_active=sql_bool(row.get("is_active"), default=True),
                created_by=sql_string(getpass.getuser()),
                notes=sql_string(row.get("notes")),
                agent_key=sql_string(row.get("agent_key")),
            )
        )
    if target_selects:
        parts.append(
            "INSERT INTO config.agent_target_month "
            "(competence_month, agent_id, target_deposit_month_usd, target_trade_day, target_trade_month, "
            "target_unique_month, target_volume_month, target_volume_unit, source_type, source_file, "
            "approved_by, is_active, created_by, notes)\n"
            + "\nUNION ALL\n".join(target_selects)
            + ";"
        )


def calendar_sql(holiday_rows: list[dict[str, Any]]) -> str:
    updates = []
    for row in holiday_rows:
        day = parse_date(row.get("date"))
        if not day:
            continue
        holiday_name = row.get("holiday_name")
        is_global_holiday = parse_bool(row.get("is_global_holiday"), default=False)
        business_override = parse_bool(row.get("is_business_day_override"), default=None)
        is_business_expr = (
            "is_business_day"
            if business_override is None
            else ("TRUE" if business_override else "FALSE")
        )
        updates.append(
            "UPDATE base_calendar SET is_global_holiday = {holiday}, holiday_name = {name}, "
            "is_business_day = {business}, source_file = {source}, notes = {notes} WHERE date = {day};".format(
                holiday=sql_bool(is_global_holiday, default=False),
                name=sql_string(holiday_name),
                business=is_business_expr,
                source=sql_string(source_file("calendar")),
                notes=sql_string(row.get("notes")),
                day=sql_string(day),
            )
        )

    override_sql = "\n".join(updates)
    return f"""
CREATE TEMP TABLE base_calendar AS
SELECT
    d::date AS date,
    TO_CHAR(d, 'YYYYMMDD')::integer AS date_sk,
    EXTRACT(YEAR FROM d)::smallint AS year,
    EXTRACT(QUARTER FROM d)::smallint AS quarter,
    EXTRACT(MONTH FROM d)::smallint AS month,
    DATE_TRUNC('month', d)::date AS month_start_date,
    EXTRACT(DAY FROM d)::smallint AS day_of_month,
    EXTRACT(ISODOW FROM d)::smallint AS day_of_week,
    TO_CHAR(d, 'FMDay') AS day_name,
    (EXTRACT(ISODOW FROM d) IN (6,7)) AS is_weekend,
    (
        TO_CHAR(d, 'MM-DD') IN ('01-01','12-25','12-26')
    ) AS is_global_holiday,
    CASE TO_CHAR(d, 'MM-DD')
        WHEN '01-01' THEN 'New Year'
        WHEN '12-25' THEN 'Christmas'
        WHEN '12-26' THEN 'Boxing Day'
        ELSE NULL
    END AS holiday_name,
    NOT (EXTRACT(ISODOW FROM d) IN (6,7) OR TO_CHAR(d, 'MM-DD') IN ('01-01','12-25','12-26')) AS is_business_day,
    NULL::varchar(255) AS source_file,
    NULL::text AS notes
FROM generate_series('2020-01-01'::date, '2035-12-31'::date, '1 day') d;

{override_sql}

INSERT INTO config.business_calendar (
    date, date_sk, year, quarter, month, month_start_date, day_of_month, day_of_week,
    day_name, is_weekend, is_global_holiday, holiday_name, is_business_day,
    business_day_number_in_month, business_days_in_month, remaining_business_days_in_month,
    source_file, notes
)
SELECT
    date,
    date_sk,
    year,
    quarter,
    month,
    month_start_date,
    day_of_month,
    day_of_week,
    day_name,
    is_weekend,
    is_global_holiday,
    holiday_name,
    is_business_day,
    CASE WHEN is_business_day THEN business_day_seq ELSE NULL END AS business_day_number_in_month,
    business_days_in_month,
    CASE WHEN is_business_day THEN business_days_in_month - business_day_seq ELSE NULL END AS remaining_business_days_in_month,
    source_file,
    notes
FROM (
    SELECT
        b.*,
        COUNT(*) FILTER (WHERE is_business_day) OVER (
            PARTITION BY month_start_date ORDER BY date ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        )::smallint AS business_day_seq,
        COUNT(*) FILTER (WHERE is_business_day) OVER (
            PARTITION BY month_start_date
        )::smallint AS business_days_in_month
    FROM base_calendar b
) x
ORDER BY date;

DROP TABLE base_calendar;
"""


def print_issues(issues: list[ValidationIssue]) -> None:
    for item in issues:
        location = f"{item.source_file}/{item.sheet_name}"
        if item.row_number:
            location += f":row {item.row_number}"
        if item.field_name:
            location += f":{item.field_name}"
        print(f"{item.severity.upper()} {location} - {item.message}")


def execute_sql(sql_path: Path, dbname: str | None, psql_args: list[str]) -> None:
    cmd = ["psql", "-v", "ON_ERROR_STOP=1"]
    if dbname:
        cmd.extend(["--dbname", dbname])
    cmd.extend(psql_args)
    cmd.extend(["--file", str(sql_path)])
    subprocess.run(cmd, check=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Load local config/domain Excel templates into PostgreSQL.")
    parser.add_argument("--template-dir", type=Path, default=DEFAULT_TEMPLATE_DIR)
    parser.add_argument("--ddl", type=Path, default=DEFAULT_DDL)
    parser.add_argument("--sql-out", type=Path, default=DEFAULT_SQL_OUT)
    parser.add_argument("--dbname", help="Database name or PostgreSQL connection string passed to psql --dbname.")
    parser.add_argument("--psql-arg", action="append", default=[], help="Extra argument passed to psql. Can be repeated.")
    parser.add_argument("--execute", action="store_true", help="Execute generated SQL with psql.")
    parser.add_argument("--check-only", action="store_true", help="Validate templates without writing SQL.")
    args = parser.parse_args()

    data = load_templates(args.template_dir)
    issues = validate(data)
    print_issues(issues)

    errors = [item for item in issues if item.severity == "error"]
    if errors:
        print(f"Validation failed with {len(errors)} error(s).", file=sys.stderr)
        return 1

    target_rows = [row for row in data["agent_target_month"] if target_has_values(row)]
    print(
        "Validated templates: "
        f"agent_profile={len(data['agent_profile'])}, "
        f"agent_alias={len(data['agent_alias'])}, "
        f"agent_target_month={len(target_rows)}, "
        f"payment_method={len(data['payment_method'])}, "
        f"lead_status={len(data['lead_status_mapping'])}, "
        f"asset_catalog={len(data['asset_catalog_override'])}."
    )

    if args.check_only:
        return 0

    sql = generate_sql(data, args.template_dir, args.ddl)
    args.sql_out.parent.mkdir(parents=True, exist_ok=True)
    args.sql_out.write_text(sql, encoding="utf-8")
    print(f"Wrote SQL: {args.sql_out}")

    if args.execute:
        execute_sql(args.sql_out, args.dbname, args.psql_arg)
        print("Load completed.")
    else:
        print("Dry run only. Re-run with --execute to apply it.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
